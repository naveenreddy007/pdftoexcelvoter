import os
import sys
import time
import logging
import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from fontTools.ttLib import TTFont
from fontTools.ttLib.tables.otTables import LigatureSubst, SingleSubst

logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)-8s | %(message)s")
logger = logging.getLogger("converter")

# ============================================================
# Table Detector (OpenCV grid-based)
# ============================================================
class TableDetector:
    def __init__(self, min_cell_area=1000, y_tolerance=15):
        self.min_cell_area = min_cell_area
        self.y_tolerance = y_tolerance

    def get_cell_matrix(self, image):
        import cv2
        if len(image.shape) == 3:
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        else:
            gray = image.copy()

        _, bin_img = cv2.threshold(gray, 200, 255, cv2.THRESH_BINARY_INV)

        kernel_len_v = gray.shape[0] // 100
        kernel_len_h = gray.shape[1] // 100

        ver_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, kernel_len_v))
        vertical_lines = cv2.erode(bin_img, ver_kernel, iterations=2)
        vertical_lines = cv2.dilate(vertical_lines, ver_kernel, iterations=3)

        hor_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (kernel_len_h, 1))
        horizontal_lines = cv2.erode(bin_img, hor_kernel, iterations=2)
        horizontal_lines = cv2.dilate(horizontal_lines, hor_kernel, iterations=3)

        table_skeleton = cv2.bitwise_or(vertical_lines, horizontal_lines)
        contours, hierarchy = cv2.findContours(table_skeleton, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

        bounding_boxes = []
        max_area = gray.shape[0] * gray.shape[1] * 0.9

        if hierarchy is None:
            return []

        for i, c in enumerate(contours):
            if hierarchy[0][i][2] != -1:
                continue
            x, y, w, h = cv2.boundingRect(c)
            area = w * h
            if self.min_cell_area < area < max_area:
                bounding_boxes.append((x, y, w, h))

        if not bounding_boxes:
            return []

        bounding_boxes.sort(key=lambda b: b[1])

        rows = []
        current_row = []
        current_y = None

        for box in bounding_boxes:
            x, y, w, h = box
            if current_y is None:
                current_y = y
                current_row.append(box)
            elif abs(y - current_y) <= self.y_tolerance:
                current_row.append(box)
                current_y = sum(b[1] for b in current_row) / len(current_row)
            else:
                current_row.sort(key=lambda b: b[0])
                rows.append(current_row)
                current_row = [box]
                current_y = y

        if current_row:
            current_row.sort(key=lambda b: b[0])
            rows.append(current_row)

        return rows

# ============================================================
# GSUB Font Rules & Decoders
# ============================================================
def build_gsub_map(font_path):
    font = TTFont(font_path)
    gsub = font['GSUB'].table
    mappings = {}
    lookups = gsub.LookupList.Lookup
    for lookup in lookups:
        for subtable in lookup.SubTable:
            if isinstance(subtable, SingleSubst):
                for input_name, output_name in subtable.mapping.items():
                    in_gid = font.getGlyphID(input_name)
                    out_gid = font.getGlyphID(output_name)
                    if out_gid not in mappings:
                        mappings[out_gid] = [in_gid]
            elif isinstance(subtable, LigatureSubst):
                for input_name, ligatures in subtable.ligatures.items():
                    in_gid = font.getGlyphID(input_name)
                    for ligature in ligatures:
                        lig_name = ligature.LigGlyph
                        lig_gid = font.getGlyphID(lig_name)
                        comp_gids = [font.getGlyphID(c) for c in ligature.Component]
                        source_seq = [in_gid] + comp_gids
                        mappings[lig_gid] = source_seq
                        
    expanded = {}
    def expand_gid(gid, path=None):
        if path is None:
            path = set()
        if gid in path:
            return [gid]
        if gid <= 115:
            return [gid]
        if gid in expanded:
            return expanded[gid]
        if gid in mappings:
            res = []
            for src in mappings[gid]:
                res.extend(expand_gid(src, path | {gid}))
            expanded[gid] = res
            return res
        else:
            return [gid]
            
    for gid in range(len(font.getGlyphOrder())):
        expand_gid(gid)
    return expanded

def reorder_telugu(text):
    vowel_signs = {'ా', 'ి', 'ీ', 'ు', 'ూ', 'ృ', 'ె', 'ే', 'ై', 'ొ', 'ో', 'ౌ'}
    chars = list(text)
    n = len(chars)
    i = 0
    while i < n - 2:
        if chars[i] in vowel_signs and chars[i+1] == '్':
            vowel = chars[i]
            chars[i] = chars[i+1]
            chars[i+1] = chars[i+2]
            chars[i+2] = vowel
            i += 2
        else:
            i += 1
    return "".join(chars)

class PDFDecoder:
    def __init__(self, font_path):
        self.expanded_map = build_gsub_map(font_path)
        font = TTFont(font_path)
        cmap = font['cmap'].getBestCmap()
        self.gid_to_unicode = {font.getGlyphID(gname): chr(u) for u, gname in cmap.items()}
        
        # Add manual mappings for unmapped GIDs in Telugu
        self.telugu_manual = {
            76: 'ఞ',
            113: 'ో',
            117: 'ి'
        }
        
        # Build manual mappings for English/numeric columns
        self.english_map = {}
        self.english_map.update({
            45: '-',
            46: '.',
            48: '0',
            49: '1',
            50: '2',
            51: '3',
            52: '4',
            53: '5',
            54: '6',
            55: '7',
            56: '8',
            57: '9',
            105: '1',  # 'ు' -> 1
            106: '3',  # 'ూ' -> 3
        })
        self.english_map.update({gid: chr(ord('A') + (gid - 65)) for gid in range(65, 91)})
        self.english_map.update({gid: chr(ord('a') + (gid - 97)) for gid in range(97, 123)})
        self.english_map.update({
            32: ' ',
            802: '-',
            805: '0',
            806: '1',
            807: '2',
            809: '4',
            811: '6',
            814: '9',
            815: ':',
            816: ';',
            818: '='
        })

    def decode_cids(self, cids, is_english=False):
        expanded = []
        for c in cids:
            if c in self.expanded_map:
                expanded.extend(self.expanded_map[c])
            else:
                expanded.append(c)
                
        decoded = []
        for c in expanded:
            if is_english:
                if c in self.english_map:
                    decoded.append(self.english_map[c])
                elif c in self.gid_to_unicode:
                    char = self.gid_to_unicode[c]
                    val_map = {
                        'ం': '3', 'అ': '5', 'ఆ': '6', 'ఇ': '7', 'ఈ': '8', 'ఉ': '9'
                    }
                    if char in val_map:
                        decoded.append(val_map[char])
                    else:
                        decoded.append(char)
                else:
                    decoded.append(f"[GID:{c}]")
            else:
                if c in self.telugu_manual:
                    decoded.append(self.telugu_manual[c])
                elif c in self.gid_to_unicode:
                    decoded.append(self.gid_to_unicode[c])
                elif c in self.english_map:
                    decoded.append(self.english_map[c])
                else:
                    decoded.append(f"[GID:{c}]")
                    
        raw_text = "".join(decoded)
        if is_english:
            return raw_text
        return reorder_telugu(raw_text)

def extract_metadata_from_page_0(page, decoder):
    spans = []
    blocks = page.get_text("dict")["blocks"]
    for b in blocks:
        if "lines" in b:
            for l in b["lines"]:
                for s in l["spans"]:
                    spans.append(s)
                    
    # Sort spans by Y (grouped by 5 points) then X
    spans.sort(key=lambda s: (round(s["origin"][1] / 5) * 5, s["origin"][0]))
    
    decoded_texts = []
    for s in spans:
        cids = [ord(c) for c in s["text"]]
        is_eng = "Gautami" not in s["font"]
        val = decoder.decode_cids(cids, is_english=is_eng).strip()
        if val:
            decoded_texts.append(val)
            
    metadata = {
        "State": "",
        "Assembly_Constituency_Name": "",
        "Assembly_Constituency_Number": "",
        "Part_Number": "",
        "Revision_Year": "",
        "Polling_Station": ""
    }
    
    for txt in decoded_texts:
        if "ఆంధ్ర ప్రదేశ్" in txt or "ఆంధ్ర ప్రదేశ్" in txt:
            metadata["State"] = "Andhra Pradesh"
            
    for i, txt in enumerate(decoded_texts):
        if "నియోజకవర్గం  సంఖ్య" in txt or "నియోజకవర్గం సంఖ్య" in txt:
            for j in range(i+1, min(i+5, len(decoded_texts))):
                cleaned = decoded_texts[j].replace(":", "").strip()
                if cleaned.isdigit():
                    metadata["Assembly_Constituency_Number"] = cleaned
                    break
                    
        if "నియోజకవర్గం  పేరు" in txt or "నియోజకవర్గం పేరు" in txt or "శాసనసభ పేరు" in txt:
            for j in range(i+1, min(i+5, len(decoded_texts))):
                cleaned = decoded_texts[j].replace(":", "").strip()
                if cleaned and not cleaned.isdigit() and len(cleaned) > 2:
                    metadata["Assembly_Constituency_Name"] = cleaned
                    break
                    
        if "భాగం సంఖ్య:" in txt or "భాగం సంఖ్య:" in txt or "భాగం సంఖ్య" in txt:
            for j in range(i+1, min(i+5, len(decoded_texts))):
                cleaned = decoded_texts[j].replace(":", "").strip()
                if cleaned.isdigit():
                    metadata["Part_Number"] = cleaned
                    break
                    
        if "సవరణ సంవత్సరం" in txt or "సవరణసంవత్సరం" in txt or "సవరణ సంవత్సరం" in txt:
            for j in range(i+1, min(i+5, len(decoded_texts))):
                cleaned = decoded_texts[j].replace(":", "").strip()
                if cleaned.isdigit() and len(cleaned) == 4:
                    metadata["Revision_Year"] = cleaned
                    break
                    
        if "చిరునామా" in txt or "పోలింగ్ కేంద్ర చిరునామా" in txt or "పోలింగ్ కేంద్ర చిరునామా" in txt:
            for j in range(i+1, min(i+5, len(decoded_texts))):
                cleaned = decoded_texts[j].replace(":", "").strip()
                if cleaned and len(cleaned) > 5:
                    metadata["Polling_Station"] = cleaned
                    break
                    
    return metadata

# ============================================================
# Main Converter Execution
# ============================================================
def convert_pdf_to_excel(pdf_path, output_path, progress_callback=None):
    start_time = time.time()
    
    if progress_callback:
        progress_callback(2, "Loading PDF...")
        
    logger.info(f"Loading PDF: {pdf_path}")
    doc = fitz.open(pdf_path)
    total_pages = doc.page_count
    logger.info(f"Total pages: {total_pages}")
    
    if progress_callback:
        progress_callback(5, "Checking fonts...")
        
    base_dir = os.path.dirname(os.path.abspath(__file__))
    font_path = os.path.join(base_dir, "ABCDEE+Gautami.ttf")
    if not os.path.exists(font_path):
        # Fallback to extract TTF font from document if missing
        logger.info("Extracting Gautami.ttf font from PDF...")
        for page_idx in range(total_pages):
            page = doc[page_idx]
            for font_info in page.get_fonts():
                xref = font_info[0]
                base_font = font_info[3]
                if "Gautami" in base_font:
                    font_name = os.path.join(base_dir, base_font.split("+")[-1] + ".ttf")
                    # Extract font tuple; the content is always the last element
                    extracted_data = doc.extract_font(xref)
                    font_bytes = extracted_data[-1] if extracted_data else None
                    if font_bytes:
                        with open(font_name, "wb") as f:
                            f.write(font_bytes)
                        font_path = font_name
                        logger.info(f"Extracted {font_path}")
                        break
            if font_path != os.path.join(base_dir, "ABCDEE+Gautami.ttf"):
                break
                
    if progress_callback:
        progress_callback(10, "Initializing PDF Decoder...")
        
    logger.info("Initializing GSUB PDF Decoder...")
    decoder = PDFDecoder(font_path)
    detector = TableDetector()
    scale_factor = 72.0 / 300
    all_rows = []
    
    if progress_callback:
        progress_callback(15, "Extracting Cover Page Metadata...")
    
    metadata = extract_metadata_from_page_0(doc[0], decoder)
    
    pages_to_process = total_pages - 2
    if pages_to_process < 1:
        pages_to_process = 1
    
    # Process all pages after cover page (summary pages are skipped automatically if no grid)
    for page_idx in range(1, total_pages):
        page_num = page_idx + 1
        page = doc[page_idx]
        
        # Render page to detect grid cells
        import numpy as np
        import cv2
        pix = page.get_pixmap(dpi=300)
        img = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.h, pix.w, pix.n)
        if pix.n == 4:
            img = cv2.cvtColor(img, cv2.COLOR_RGBA2BGR)
        elif pix.n == 3:
            img = cv2.cvtColor(img, cv2.COLOR_RGB2BGR)
            
        cv_rows = detector.get_cell_matrix(img)
        data_rows = [r for r in cv_rows if len(r) >= 6]
        
        if not data_rows:
            logger.warning(f"Page {page_num}: No data table detected, skipping.")
            continue
            
        # Get raw text spans
        raw_spans = []
        blocks = page.get_text("dict")["blocks"]
        for b in blocks:
            if "lines" in b:
                for l in b["lines"]:
                    for s in l["spans"]:
                        raw_spans.append(s)
                        
        page_rows = []
        # Skip header rows (index 0 and 1)
        for r_idx, row in enumerate(data_rows[2:]):
            row_text = []
            for col_idx, (x, y, w, h) in enumerate(row):
                if col_idx >= 8:
                    break
                cell_x0, cell_y0 = x * scale_factor, y * scale_factor
                cell_x1, cell_y1 = (x + w) * scale_factor, (y + h) * scale_factor
                
                cell_spans = []
                for s in raw_spans:
                    sx, sy = s["origin"][0], s["origin"][1]
                    if cell_x0 - 2 <= sx <= cell_x1 + 2 and cell_y0 - 2 <= sy <= cell_y1 + 2:
                        cell_spans.append(s)
                
                raw_text = "".join(s["text"] for s in cell_spans)
                cids = [ord(c) for c in raw_text]
                
                is_eng = (col_idx in [0, 1, 6, 7])
                decoded_val = decoder.decode_cids(cids, is_english=is_eng).strip()
                
                # Special cleanup for relation and gender
                if col_idx == 3:
                    clean = decoded_val.replace(" ", "")
                    if "తం" in clean:
                        decoded_val = "తం"
                    elif "భ" in clean:
                        decoded_val = "భ"
                elif col_idx == 5:
                    clean = decoded_val.replace(" ", "")
                    if "పు" in clean or "ప" in clean:
                        decoded_val = "పు"
                    elif "స్త్ర" in clean or "స్త" in clean or "స" in clean:
                        decoded_val = "స్త్రీ"
                        
                row_text.append(decoded_val)
                
            if len(row_text) >= 2:
                while len(row_text) < 8:
                    row_text.append("")
                page_rows.append(row_text[:8])
                
        all_rows.extend(page_rows)
        logger.info(f"Page {page_num}: Extracted {len(page_rows)} rows. Running total: {len(all_rows)}")
        
        if progress_callback:
            percent = 15 + int(((page_idx) / pages_to_process) * 75)
            progress_callback(percent, f"Processed page {page_num} of {total_pages}...")
            
    doc.close()
    
    if progress_callback:
        progress_callback(90, "Generating formatted Excel sheets...")
    
    # Write final Excel
    logger.info(f"Writing Excel with {len(all_rows)} rows...")
    wb = Workbook()
    
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    data_font = Font(name="Arial", size=10)
    data_alignment = Alignment(vertical="center", wrap_text=True)
    
    # 1. Write Voters Sheet FIRST (so it opens by default)
    ws = wb.active
    ws.title = "Voters"
    
    COLUMN_HEADERS = [
        "వరుస సంఖ్య (Sl No)",
        "ఇంటి నంబర్ (House No)",
        "ఓటరు పేరు (Voter Name)",
        "సంబంధం (Relation)",
        "తండ్రి/భర్త పేరు (Father/Husband)",
        "లింగం (Gender)",
        "వయసు (Age)",
        "ఓటరు కార్డు నంబర్ (Voter ID)"
    ]
    
    for col_idx, header in enumerate(COLUMN_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
    for row_idx, row_data in enumerate(all_rows, start=2):
        for col_idx, cell_value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            
    col_widths = [12, 16, 32, 12, 32, 12, 10, 24]
    for idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
        
    ws.freeze_panes = "A2"
    
    # 2. Write Metadata Sheet (second tab)
    ws_meta = wb.create_sheet(title="Metadata")
    
    metadata_headers = ["State", "Assembly_Constituency_Name", "Assembly_Constituency_Number", "Part_Number", "Revision_Year", "Polling_Station", "Total_Voters_Extracted"]
    metadata["Total_Voters_Extracted"] = str(len(all_rows))
    for col_idx, header in enumerate(metadata_headers, start=1):
        cell = ws_meta.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
    for col_idx, header in enumerate(metadata_headers, start=1):
        val = metadata.get(header, "")
        cell = ws_meta.cell(row=2, column=col_idx, value=val)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
        
    for col_idx in range(1, len(metadata_headers) + 1):
        ws_meta.column_dimensions[get_column_letter(col_idx)].width = 25
    
    # Make sure Voters sheet is the active one when file opens
    wb.active = 0
    
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    wb.save(output_path)
    
    if progress_callback:
        progress_callback(100, "Extraction complete!")
        
    elapsed = time.time() - start_time
    file_size = os.path.getsize(output_path)
    
    logger.info("=" * 60)
    logger.info("CONVERSION COMPLETE")
    logger.info(f"  Input:  {pdf_path}")
    logger.info(f"  Output: {output_path}")
    logger.info(f"  Pages:  {total_pages}")
    logger.info(f"  Rows:   {len(all_rows)}")
    logger.info(f"  Size:   {file_size / 1024:.1f} KB")
    logger.info(f"  Time:   {elapsed:.1f}s")
    logger.info("=" * 60)
    
    return output_path, len(all_rows)

if __name__ == "__main__":
    pdf_file = sys.argv[1] if len(sys.argv) > 1 else "DOC-20260613-WA0000..pdf"
    output_file = sys.argv[2] if len(sys.argv) > 2 else "voter_list_output.xlsx"
    convert_pdf_to_excel(pdf_file, output_file)
